from django.core.management.base import BaseCommand
from django.db import transaction

from homepage.models import SopAnswer, SopGameTest, SopQuestion
from homepage.sop_import_config import SOP_IMPORT_SOURCES, SOP_XLSX_DIR


def _cell_str(val):
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


class Command(BaseCommand):
    help = 'Импорт тестов СОП из Excel в static/boomerang/sop/ (см. sop_import_config).'

    def add_arguments(self, parser):
        parser.add_argument(
            '--replace',
            action='store_true',
            help='Удалить существующие вопросы по тем же slug и загрузить заново',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as e:
            self.stderr.write('Нужен пакет openpyxl: pip install openpyxl')
            raise e

        replace = options['replace']
        total_q = 0

        for spec in SOP_IMPORT_SOURCES:
            path = SOP_XLSX_DIR / spec['file']
            if not path.is_file():
                self.stderr.write(self.style.WARNING(f'Пропуск — нет файла: {path}'))
                continue

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet_cfg in spec['sheets']:
                    slug = sheet_cfg['slug']
                    title = sheet_cfg['title']
                    sh = sheet_cfg['sheet']
                    if isinstance(sh, int):
                        ws = wb.worksheets[sh]
                    else:
                        if sh not in wb.sheetnames:
                            self.stderr.write(
                                self.style.WARNING(f'Лист «{sh}» не найден в {spec["file"]}')
                            )
                            continue
                        ws = wb[sh]

                    n_imported = self._import_sheet(
                        ws, slug=slug, title=title, replace=replace
                    )
                    total_q += n_imported
                    self.stdout.write(
                        f'{slug}: импортировано вопросов: {n_imported} ({title})'
                    )
            finally:
                wb.close()

        self.stdout.write(self.style.SUCCESS(f'Готово. Всего новых/обновлённых вопросов: {total_q}'))

    @transaction.atomic
    def _import_sheet(self, ws, *, slug, title, replace):
        test, _ = SopGameTest.objects.get_or_create(
            slug=slug,
            defaults={'title': title, 'description': ''},
        )
        if test.title != title:
            test.title = title
            test.save(update_fields=['title', 'updated_at'])

        if test.questions.exists() and not replace:
            self.stderr.write(
                self.style.WARNING(
                    f'{slug}: в тесте уже есть вопросы — пропуск '
                    f'(используйте --replace для перезаписи)'
                )
            )
            return 0

        if replace:
            test.questions.all().delete()

        sort_order = 0
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 6:
                continue
            q_text = _cell_str(row[1])
            correct = _cell_str(row[2])
            wrong = [_cell_str(row[i]) for i in (3, 4, 5)]
            wrong_cells = [w for w in wrong if w]
            if not q_text or not correct:
                continue
            if not wrong_cells:
                continue

            sort_order += 1
            question = SopQuestion.objects.create(
                test=test,
                text=q_text,
                sort_order=sort_order,
            )
            SopAnswer.objects.create(
                question=question,
                text=correct,
                is_correct=True,
            )
            for w in wrong_cells:
                SopAnswer.objects.create(
                    question=question,
                    text=w,
                    is_correct=False,
                )
            n += 1
        return n
